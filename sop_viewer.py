import ctypes
import io
import os
import re
import sys
import zipfile
import customtkinter as ctk
from openpyxl import load_workbook
import pandas as pd
from PIL import Image

# Configuración del tema visual
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

EXCEL_FILE = "Prueba1.xlsx"


# ==============================================================================
# FUNCIONES AUXILIARES Y PROCESAMIENTO DE IMÁGENES
# ==============================================================================
def ocultar_archivo_windows(ruta):
  """Oculta un archivo o carpeta en sistemas Windows mediante la API Win32"""
  if os.name == "nt" and os.path.exists(ruta):
    try:
      ctypes.windll.kernel32.SetFileAttributesW(str(ruta), 2)
    except Exception:
      pass


def limpiar_texto(val):
  if pd.isna(val):
    return ""
  s = str(val).strip()
  return "" if s.lower() in ["nan", "none"] else s


def limpiar_entero(val):
  """Convierte flotantes o textos con .0 a entero limpio (ej. '90.0' -> '90')"""
  s = limpiar_texto(val)
  if not s:
    return ""
  try:
    num = float(s)
    if num.is_integer():
      return str(int(num))
    return str(num)
  except ValueError:
    return s


def normalizar_texto(texto):
  texto = limpiar_texto(texto).lower()
  reemplazos = (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"))
  for a, b in reemplazos:
    texto = texto.replace(a, b)
  return texto


def normalizar_fondo_blanco(pil_img):
  """Convierte imágenes con transparencia (RGBA/PNG) o recorte a un lienzo blanco sólido"""
  try:
    if pil_img.mode in ("RGBA", "LA") or (
        pil_img.mode == "P" and "transparency" in pil_img.info
    ):
      alpha_img = pil_img.convert("RGBA")
      canvas = Image.new("RGBA", alpha_img.size, (255, 255, 255, 255))
      canvas.paste(alpha_img, mask=alpha_img.split()[3])
      return canvas.convert("RGB")
    return pil_img.convert("RGB")
  except Exception:
    return pil_img.convert("RGB")


# ==============================================================================
# MOTOR ROBUSTO DE EXTRACCIÓN DE IMÁGENES DE EXCEL
# ==============================================================================
def extraer_imagenes_hoja(ruta_excel, nombre_hoja, col_mat_idx=0):
  """Extrae imágenes de la hoja de Excel aplicando normalización de fondo blanco"""
  mapa_imgs = {}
  if not os.path.exists(ruta_excel):
    return mapa_imgs

  try:
    wb = load_workbook(ruta_excel, data_only=True)
    if nombre_hoja not in wb.sheetnames:
      return mapa_imgs
    ws = wb[nombre_hoja]

    filas_mat = {}
    for row in range(2, ws.max_row + 1):
      val_mat = limpiar_texto(ws.cell(row=row, column=col_mat_idx + 1).value)
      if val_mat:
        filas_mat[row] = val_mat

    # Método A: Extracción vía openpyxl _images
    for img in ws._images:
      try:
        row_target = img.anchor._from.row + 1
        if row_target in filas_mat:
          mat = filas_mat[row_target]
          pil_img = Image.open(io.BytesIO(img._data()))
          pil_img = normalizar_fondo_blanco(pil_img)
          if mat not in mapa_imgs:
            mapa_imgs[mat] = []
          mapa_imgs[mat].append(pil_img)
      except Exception:
        pass

    # Método B: Extracción vía ZIP/XML para celdas incrustadas (Place in cell)
    if not mapa_imgs:
      with zipfile.ZipFile(ruta_excel, "r") as z:
        media_files = [
            f for f in z.namelist() if f.startswith("xl/media/")
        ]
        if media_files:
          materiales_ordenados = list(dict.fromkeys(filas_mat.values()))
          for idx, media_path in enumerate(sorted(media_files)):
            if idx < len(materiales_ordenados):
              mat = materiales_ordenados[idx]
              img_data = z.read(media_path)
              pil_img = Image.open(io.BytesIO(img_data))
              pil_img = normalizar_fondo_blanco(pil_img)
              if mat not in mapa_imgs:
                mapa_imgs[mat] = []
              mapa_imgs[mat].append(pil_img)

  except Exception as e:
    print(f"[AVISO] Error al procesar imágenes de {nombre_hoja}: {e}")

  return mapa_imgs


# ==============================================================================
# APLICACIÓN PRINCIPAL
# ==============================================================================
class SOPApp(ctk.CTk):

  def __init__(self):
    super().__init__()

    ruta_internal = os.path.join(os.path.dirname(sys.executable), "_internal")
    if os.path.exists(ruta_internal):
      ocultar_archivo_windows(ruta_internal)

    self.title("QRs Componentes - Visor de SOPs")
    self.geometry("1380x880")
    self.minsize(1150, 750)

    # DataFrames de Excel
    self.data_general = pd.DataFrame()
    self.data_estandar = pd.DataFrame()
    self.data_componentes = pd.DataFrame()
    self.data_alertas = pd.DataFrame()
    self.data_pasos = pd.DataFrame()
    self.data_puntos_seg = pd.DataFrame()
    self.data_epp_ob = pd.DataFrame()
    self.data_epp_req = pd.DataFrame()
    self.data_historial = pd.DataFrame()

    # Cachés de imágenes procesadas con fondo blanco
    self.imgs_general = {}
    self.imgs_componentes = {}
    self.imgs_pasos = {}
    self.imgs_epp_ob = {}
    self.imgs_epp_req = {}

    self.materiales_unicos = []
    self.material_actual = None
    self.maquina_actual = None
    self.paso_actual_index = 0
    self.pasos_filtrados = []

    self._crear_interfaz()
    self.cargar_datos_excel()

  def _crear_interfaz(self):
    # 1. Header Buscador
    self.frame_header = ctk.CTkFrame(self, corner_radius=10)
    self.frame_header.pack(fill="x", padx=15, pady=10)

    self.lbl_titulo_app = ctk.CTkLabel(
        self.frame_header,
        text="📘 DIGITALIZACIÓN DE SOPs",
        font=("Helvetica", 16, "bold"),
        text_color="#1F4E79",
    )
    self.lbl_titulo_app.pack(side="left", padx=15, pady=10)

    self.frame_search_container = ctk.CTkFrame(
        self.frame_header, fg_color="transparent"
    )
    self.frame_search_container.pack(
        side="right", fill="x", expand=True, padx=15, pady=10
    )

    self.entry_busqueda = ctk.CTkEntry(
        self.frame_search_container,
        placeholder_text="🔍 Buscar por Material o Máquina...",
        font=("Helvetica", 13),
        height=38,
    )
    self.entry_busqueda.pack(fill="x", expand=True)
    self.entry_busqueda.bind("<KeyRelease>", self.al_escribir_buscador)

    self.frame_sugerencias = ctk.CTkScrollableFrame(
        self.frame_search_container,
        height=150,
        corner_radius=6,
        border_width=1,
        border_color="#1F4E79",
    )

    # 2. Banner Superior
    self.frame_banner = ctk.CTkFrame(self, corner_radius=8, fg_color="#E6F0FA")
    self.frame_banner.pack(fill="x", padx=15, pady=(0, 10))

    self.lbl_banner_mat = ctk.CTkLabel(
        self.frame_banner,
        text="MATERIAL: ---",
        font=("Helvetica", 17, "bold"),
        text_color="#111111",
    )
    self.lbl_banner_mat.pack(side="left", padx=15, pady=8)

    self.lbl_banner_maq = ctk.CTkLabel(
        self.frame_banner,
        text="MÁQUINA: ---",
        font=("Helvetica", 17, "bold"),
        text_color="#1F4E79",
    )
    self.lbl_banner_maq.pack(side="left", padx=15, pady=8)

    self.lbl_banner_desc = ctk.CTkLabel(
        self.frame_banner,
        text="",
        font=("Helvetica", 15, "bold", "italic"),
        text_color="#222222",
    )
    self.lbl_banner_desc.pack(side="right", padx=15, pady=8)

    # 3. Pestañas
    self.tabview = ctk.CTkTabview(self, corner_radius=10)
    self.tabview.pack(fill="both", expand=True, padx=15, pady=(0, 10))

    self.tab_info = self.tabview.add("1. Info & Estándar")
    self.tab_comp = self.tabview.add("2. Componentes")
    self.tab_seg = self.tabview.add("3. Seguridad & EPP")
    self.tab_alertas = self.tabview.add("4. Alertas & Docs")
    self.tab_pasos = self.tabview.add("5. Pasos Operativos")
    self.tab_historial = self.tabview.add("6. Historial Cambios")

    self._setup_tab_pasos()

  # ==============================================================================
  # CARGA DE DATOS
  # ==============================================================================
  def cargar_datos_excel(self):
    if not os.path.exists(EXCEL_FILE):
      return

    try:
      xls = pd.ExcelFile(EXCEL_FILE)
      self.data_general = pd.read_excel(xls, sheet_name=0)
      self.data_estandar = pd.read_excel(xls, sheet_name=1)
      self.data_componentes = pd.read_excel(xls, sheet_name=2)
      self.data_alertas = pd.read_excel(xls, sheet_name=3)
      self.data_pasos = pd.read_excel(xls, sheet_name=4)
      self.data_puntos_seg = pd.read_excel(xls, sheet_name=5)
      self.data_epp_ob = pd.read_excel(xls, sheet_name=6)
      self.data_epp_req = pd.read_excel(xls, sheet_name=7)
      self.data_historial = pd.read_excel(xls, sheet_name=8)

      self.imgs_general = extraer_imagenes_hoja(EXCEL_FILE, xls.sheet_names[0])
      self.imgs_componentes = extraer_imagenes_hoja(
          EXCEL_FILE, xls.sheet_names[2]
      )
      self.imgs_pasos = extraer_imagenes_hoja(EXCEL_FILE, xls.sheet_names[4])
      self.imgs_epp_ob = extraer_imagenes_hoja(EXCEL_FILE, xls.sheet_names[6])
      self.imgs_epp_req = extraer_imagenes_hoja(EXCEL_FILE, xls.sheet_names[7])

      col_mat = self.data_general.columns[0]
      col_maq = (
          self.data_general.columns[2]
          if len(self.data_general.columns) > 2
          else col_mat
      )
      col_desc = (
          self.data_general.columns[6]
          if len(self.data_general.columns) > 6
          else col_mat
      )

      self.materiales_unicos.clear()
      for _, row in self.data_general.iterrows():
        m = limpiar_texto(row[col_mat])
        mq = limpiar_texto(row[col_maq])
        d = limpiar_texto(row[col_desc])
        if m:
          self.materiales_unicos.append((m, mq, d))

      if self.materiales_unicos:
        pm, pmq, _ = self.materiales_unicos[0]
        self.seleccionar_material(pm, pmq)

    except Exception as e:
      print(f"Error al procesar Excel: {e}")

  # ==============================================================================
  # BUSCADOR
  # ==============================================================================
  def al_escribir_buscador(self, event=None):
    query = normalizar_texto(self.entry_busqueda.get())
    if not query or not self.materiales_unicos:
      self.frame_sugerencias.pack_forget()
      return

    for child in self.frame_sugerencias.winfo_children():
      child.destroy()

    coincidencias = 0
    for mat, mq, desc in self.materiales_unicos:
      eval_txt = f"{mat} {mq} {desc}"
      if query in normalizar_texto(eval_txt):
        lbl_btn = f"{mat}  │  Máq: {mq}" + (f" ({desc})" if desc else "")
        btn = ctk.CTkButton(
            self.frame_sugerencias,
            text=lbl_btn,
            anchor="w",
            fg_color="transparent",
            text_color="#111111",
            hover_color="#D0E0F0",
            height=28,
            command=lambda m=mat, mq=mq: self.seleccionar_material(m, mq),
        )
        btn.pack(fill="x", pady=1)
        coincidencias += 1
        if coincidencias >= 15:
          break

    if coincidencias > 0:
      self.frame_sugerencias.pack(fill="x", pady=(4, 0))
    else:
      self.frame_sugerencias.pack_forget()

  def seleccionar_material(self, material, maquina):
    self.frame_sugerencias.pack_forget()
    self.entry_busqueda.delete(0, "end")
    self.focus_set()

    self.material_actual = material
    self.maquina_actual = maquina

    self.lbl_banner_mat.configure(text=f"MATERIAL: {material}")
    self.lbl_banner_maq.configure(text=f"MÁQUINA: {maquina}")

    self.renderizar_todas_pestañas(material, maquina)

  def renderizar_todas_pestañas(self, material, maquina):
    self.renderizar_tab_info(material, maquina)
    self.renderizar_tab_componentes(material)
    self.renderizar_tab_seguridad(material)
    self.renderizar_tab_alertas(material)
    self.renderizar_tab_pasos(material)
    self.renderizar_tab_historial(material)

  # ==============================================================================
  # PESTAÑA 1: INFO & ESTÁNDAR (FORMATO EQUILIBRADO SIN ESPACIO VACÍO ABAJO)
  # ==============================================================================
  def renderizar_tab_info(self, material, maquina):
    for child in self.tab_info.winfo_children():
      child.destroy()

    frame_main = ctk.CTkFrame(self.tab_info, fg_color="transparent")
    frame_main.pack(fill="both", expand=True, padx=10, pady=5)

    # COLUMNA 1: Datos Generales
    col_izq = ctk.CTkFrame(frame_main, fg_color="transparent")
    col_izq.pack(side="left", fill="both", expand=False, padx=(0, 10))

    # COLUMNA 2: Estándar y Herramientas
    col_centro = ctk.CTkFrame(frame_main, fg_color="transparent")
    col_centro.pack(side="left", fill="both", expand=False, padx=(0, 10))

    # COLUMNA 3: Imagen Principal
    col_der = ctk.CTkFrame(
        frame_main,
        width=480,
        corner_radius=10,
        border_width=1,
        border_color="#CCCCCC",
        fg_color="#FFFFFF",
    )
    col_der.pack(side="right", fill="both", expand=True)

    # 1. DATOS GENERALES
    col_mat_g = self.data_general.columns[0]
    df_g = self.data_general[
        self.data_general[col_mat_g].apply(limpiar_texto) == material
    ]

    if not df_g.empty:
      row = df_g.iloc[0]
      desc = limpiar_texto(row.iloc[6]) if len(row) > 6 else ""
      self.lbl_banner_desc.configure(text=desc)

      ctk.CTkLabel(
          col_izq,
          text="📋 DATOS GENERALES",
          font=("Helvetica", 16, "bold"),
          text_color="#1F4E79",
      ).pack(anchor="w", pady=(0, 4))

      tbl_g = ctk.CTkFrame(
          col_izq,
          fg_color="#B0C4DE",
          border_width=1,
          border_color="#1F4E79",
          corner_radius=4,
      )
      tbl_g.pack(fill="both", expand=True, pady=(0, 2))

      tbl_g.grid_columnconfigure(0, weight=1)
      tbl_g.grid_columnconfigure(1, weight=2)

      cols = self.data_general.columns
      f_idx = 0
      for col_name in cols:
        val = limpiar_texto(row[col_name])
        if col_name.lower().startswith("imagen"):
          continue

        nombre_limpio = str(col_name).strip().rstrip(":")

        c_k = ctk.CTkFrame(tbl_g, fg_color="#F0F4F8", corner_radius=0)
        c_k.grid(row=f_idx, column=0, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(
            c_k,
            text=f"{nombre_limpio}:",
            font=("Helvetica", 12, "bold"),
            text_color="#1F4E79",
            anchor="w",
        ).pack(anchor="w", padx=6, pady=2)

        c_v = ctk.CTkFrame(tbl_g, fg_color="#FFFFFF", corner_radius=0)
        c_v.grid(row=f_idx, column=1, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(
            c_v,
            text=val if val else "---",
            font=("Helvetica", 13, "bold"),
            text_color="#000000",
            anchor="w",
        ).pack(anchor="w", padx=6, pady=2)

        f_idx += 1

    # 2. ESTÁNDAR DE PRODUCCIÓN Y HERRAMIENTAS
    col_mat_e = self.data_estandar.columns[0]
    df_e = self.data_estandar[
        self.data_estandar[col_mat_e].apply(limpiar_texto) == material
    ]

    if not df_e.empty:
      ctk.CTkLabel(
          col_centro,
          text="⚙️ ESTÁNDAR DE PRODUCCIÓN",
          font=("Helvetica", 16, "bold"),
          text_color="#1F4E79",
      ).pack(anchor="w", pady=(0, 4))

      row1 = df_e.iloc[0]

      pzs_hr = limpiar_entero(row1.iloc[1])
      t1 = limpiar_entero(row1.iloc[2])
      t2 = limpiar_entero(row1.iloc[3])
      t3 = limpiar_entero(row1.iloc[4])
      pzs_ciclo = limpiar_entero(row1.iloc[5])
      tiempo_ciclo = limpiar_texto(row1.iloc[6])
      if tiempo_ciclo and not tiempo_ciclo.lower().endswith("seg."):
        tiempo_ciclo = f"{limpiar_entero(tiempo_ciclo)} seg."
      wip_max = limpiar_entero(row1.iloc[7])

      tbl = ctk.CTkFrame(
          col_centro,
          fg_color="#000000",
          border_width=1,
          border_color="#000000",
          corner_radius=0,
      )
      tbl.pack(fill="x", pady=(0, 10))

      tbl.grid_columnconfigure(0, weight=1)
      tbl.grid_columnconfigure(1, weight=1)
      tbl.grid_columnconfigure(2, weight=1)
      tbl.grid_columnconfigure(3, weight=1)

      c_est = ctk.CTkFrame(tbl, fg_color="#FFFFFF", corner_radius=0)
      c_est.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)
      ctk.CTkLabel(
          c_est,
          text="Estándar\n(Pzs/Hr.)",
          font=("Helvetica", 12, "bold"),
          text_color="#000000",
      ).pack(expand=True, pady=2, padx=4)

      c_turnos_tit = ctk.CTkFrame(tbl, fg_color="#FFFFFF", corner_radius=0)
      c_turnos_tit.grid(
          row=0, column=1, columnspan=3, sticky="nsew", padx=1, pady=1
      )
      ctk.CTkLabel(
          c_turnos_tit,
          text="Piezas por turno (al 100%)",
          font=("Helvetica", 13, "bold"),
          text_color="#000000",
      ).pack(pady=2)

      headers_turnos = ["1er Turno", "2do. Turno", "3er Turno."]
      for idx_t, ht in enumerate(headers_turnos):
        c_h = ctk.CTkFrame(tbl, fg_color="#FFFFFF", corner_radius=0)
        c_h.grid(row=1, column=idx_t + 1, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(
            c_h,
            text=ht,
            font=("Helvetica", 12, "bold"),
            text_color="#000000",
        ).pack(pady=2, padx=2)

      valores_f3 = [pzs_hr, t1, t2, t3]
      for idx_v, val_num in enumerate(valores_f3):
        c_v = ctk.CTkFrame(tbl, fg_color="#FFFFFF", corner_radius=0)
        c_v.grid(row=2, column=idx_v, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(
            c_v,
            text=val_num,
            font=("Helvetica", 24, "bold"),
            text_color="#000000",
        ).pack(pady=4, padx=2)

      filas_inferiores = [
          ("PIEZAS POR CICLO:", pzs_ciclo),
          ("TIEMPO CICLO:", tiempo_ciclo),
          ("WIP MAX:", wip_max),
      ]

      for idx_f, (k_inf, v_inf) in enumerate(filas_inferiores):
        r_idx = 3 + idx_f

        c_lbl = ctk.CTkFrame(tbl, fg_color="#FFFFFF", corner_radius=0)
        c_lbl.grid(row=r_idx, column=0, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(
            c_lbl,
            text=k_inf,
            font=("Helvetica", 12, "bold"),
            text_color="#000000",
            anchor="e",
        ).pack(side="right", padx=4, pady=3)

        c_val = ctk.CTkFrame(tbl, fg_color="#FFFFFF", corner_radius=0)
        c_val.grid(
            row=r_idx, column=1, columnspan=3, sticky="nsew", padx=1, pady=1
        )
        ctk.CTkLabel(
            c_val,
            text=v_inf,
            font=("Helvetica", 15, "bold"),
            text_color="#000000",
        ).pack(expand=True, pady=3)

      col_herram = df_e.columns[8] if len(df_e.columns) > 8 else None
      if col_herram:
        herramientas = [
            limpiar_texto(h) for h in df_e[col_herram].dropna() if limpiar_texto(h)
        ]
        if herramientas:
          ctk.CTkLabel(
              col_centro,
              text="🔧 HERRAMIENTAS REQUERIDAS",
              font=("Helvetica", 16, "bold"),
              text_color="#1F4E79",
          ).pack(anchor="w", pady=(6, 4))

          frame_herramientas_box = ctk.CTkFrame(
              col_centro, fg_color="transparent"
          )
          frame_herramientas_box.pack(fill="both", expand=True)

          for h_item in herramientas:
            card_h = ctk.CTkFrame(
                frame_herramientas_box,
                fg_color="#FFFFFF",
                border_width=1,
                border_color="#D1D5DB",
                corner_radius=5,
            )
            card_h.pack(fill="x", pady=2)
            ctk.CTkLabel(
                card_h,
                text=f"• {h_item}",
                font=("Helvetica", 13, "bold"),
                text_color="#222222",
                anchor="w",
            ).pack(anchor="w", padx=10, pady=5)

    # 3. FOTOGRAFÍA PRINCIPAL
    imgs_m = self.imgs_general.get(material, [])
    if imgs_m:
      pil_img = imgs_m[0]
      ctk_img = ctk.CTkImage(
          light_image=pil_img, dark_image=pil_img, size=(450, 450)
      )
      lbl_img = ctk.CTkLabel(col_der, image=ctk_img, text="")
      lbl_img.pack(expand=True, fill="both", padx=10, pady=10)
    else:
      ctk.CTkLabel(
          col_der,
          text="📷 Fotografía del Material\n(No disponible)",
          font=("Helvetica", 15, "bold"),
          text_color="#888888",
      ).pack(expand=True)

  # ==============================================================================
  # PESTAÑA 2: COMPONENTES
  # ==============================================================================
  def renderizar_tab_componentes(self, material):
    for child in self.tab_comp.winfo_children():
      child.destroy()

    scroll = ctk.CTkScrollableFrame(self.tab_comp)
    scroll.pack(fill="both", expand=True)

    col_mat_c = self.data_componentes.columns[0]
    df_c = self.data_componentes[
        self.data_componentes[col_mat_c].apply(limpiar_texto) == material
    ]

    if df_c.empty:
      ctk.CTkLabel(
          scroll,
          text="No hay componentes registrados para este material.",
          font=("Helvetica", 15, "bold"),
      ).pack(pady=40)
      return

    scroll.grid_columnconfigure(0, weight=1)
    scroll.grid_columnconfigure(1, weight=1)

    imgs_mat = self.imgs_componentes.get(material, [])

    for idx, (_, row) in enumerate(df_c.iterrows()):
      nombre = limpiar_texto(row.iloc[1]) if len(row) > 1 else ""
      no_parte = limpiar_texto(row.iloc[2]) if len(row) > 2 else ""
      codigo = limpiar_texto(row.iloc[3]) if len(row) > 3 else ""

      r, c = divmod(idx, 2)
      card = ctk.CTkFrame(
          scroll,
          corner_radius=8,
          border_width=1,
          border_color="#CCCCCC",
          fg_color="#FFFFFF",
      )
      card.grid(row=r, column=c, padx=10, pady=10, sticky="nsew")

      lbl_cod = ctk.CTkLabel(
          card,
          text=codigo if codigo else nombre,
          font=("Helvetica", 16, "bold"),
          text_color="#1F4E79",
      )
      lbl_cod.pack(anchor="w", padx=14, pady=(12, 2))

      if no_parte:
        ctk.CTkLabel(
            card,
            text=f"No. Parte: {no_parte}",
            font=("Helvetica", 14, "bold"),
            text_color="#555555",
        ).pack(anchor="w", padx=14, pady=2)
      if nombre and codigo:
        ctk.CTkLabel(
            card,
            text=nombre,
            font=("Helvetica", 14),
            text_color="#222222",
        ).pack(anchor="w", padx=14, pady=2)

      if idx < len(imgs_mat):
        pil_img = imgs_mat[idx]
        ctk_img = ctk.CTkImage(
            light_image=pil_img, dark_image=pil_img, size=(150, 150)
        )
        ctk.CTkLabel(card, image=ctk_img, text="").pack(pady=10)

  # ==============================================================================
  # PESTAÑA 3: SEGURIDAD & EPP
  # ==============================================================================
  def renderizar_tab_seguridad(self, material):
    for child in self.tab_seg.winfo_children():
      child.destroy()

    scroll = ctk.CTkScrollableFrame(self.tab_seg)
    scroll.pack(fill="both", expand=True)

    frame_ob = ctk.CTkFrame(scroll, corner_radius=8)
    frame_ob.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        frame_ob,
        text="🛑 EQUIPO DE SEGURIDAD OBLIGATORIO",
        font=("Helvetica", 16, "bold"),
        text_color="#D32F2F",
    ).pack(anchor="w", padx=15, pady=12)

    imgs_ob = self.imgs_epp_ob.get(material, [])
    if imgs_ob:
      grid_ob = ctk.CTkFrame(frame_ob, fg_color="transparent")
      grid_ob.pack(fill="x", padx=15, pady=(0, 10))

      for pil_img in imgs_ob:
        ctk_img = ctk.CTkImage(
            light_image=pil_img, dark_image=pil_img, size=(110, 110)
        )
        card_i = ctk.CTkFrame(
            grid_ob,
            fg_color="#FFFFFF",
            border_width=1,
            border_color="#E0E0E0",
            corner_radius=6,
        )
        card_i.pack(side="left", padx=8, pady=8)
        ctk.CTkLabel(card_i, image=ctk_img, text="").pack(padx=8, pady=8)

    frame_req = ctk.CTkFrame(scroll, corner_radius=8)
    frame_req.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        frame_req,
        text="⚠️ EQUIPO ESPECIAL (SI SE REQUIERE)",
        font=("Helvetica", 16, "bold"),
        text_color="#1F4E79",
    ).pack(anchor="w", padx=15, pady=12)

    col_mat_req = self.data_epp_req.columns[0]
    df_req = self.data_epp_req[
        self.data_epp_req[col_mat_req].apply(limpiar_texto) == material
    ]

    imgs_req = self.imgs_epp_req.get(material, [])

    if not df_req.empty or imgs_req:
      grid_req = ctk.CTkFrame(frame_req, fg_color="transparent")
      grid_req.pack(fill="x", padx=15, pady=(0, 10))

      for idx in range(max(len(df_req), len(imgs_req))):
        requiere = "NO"
        if idx < len(df_req):
          row = df_req.iloc[idx]
          requiere = (
              limpiar_texto(row.iloc[2]).upper() if len(row) > 2 else "NO"
          )

        es_si = "SI" in requiere

        card_r = ctk.CTkFrame(
            grid_req,
            fg_color="#E8F5E9" if es_si else "#F5F5F5",
            border_width=2 if es_si else 1,
            border_color="#2E7D32" if es_si else "#CCCCCC",
            corner_radius=6,
        )
        card_r.pack(side="left", padx=8, pady=8)

        lbl_st = f"✓ REQUIERE" if es_si else "NO REQUIERE"
        ctk.CTkLabel(
            card_r,
            text=lbl_st,
            font=("Helvetica", 13, "bold"),
            text_color="#2E7D32" if es_si else "#888888",
        ).pack(pady=(8, 4), padx=10)

        if idx < len(imgs_req):
          pil_img = imgs_req[idx]
          ctk_img = ctk.CTkImage(
              light_image=pil_img, dark_image=pil_img, size=(100, 100)
          )
          ctk.CTkLabel(card_r, image=ctk_img, text="").pack(padx=10, pady=(0, 8))

  # ==============================================================================
  # PESTAÑA 4: ALERTAS & DOCS
  # ==============================================================================
  def renderizar_tab_alertas(self, material):
    for child in self.tab_alertas.winfo_children():
      child.destroy()

    scroll = ctk.CTkScrollableFrame(self.tab_alertas)
    scroll.pack(fill="both", expand=True)

    col_mat_a = self.data_alertas.columns[0]
    df_a = self.data_alertas[
        self.data_alertas[col_mat_a].apply(limpiar_texto) == material
    ]

    if df_a.empty:
      ctk.CTkLabel(
          scroll,
          text="No hay alertas ni documentos registrados.",
          font=("Helvetica", 15, "bold"),
      ).pack(pady=40)
      return

    frame_a = ctk.CTkFrame(scroll, corner_radius=8)
    frame_a.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        frame_a,
        text="🚨 ALERTAS DE CALIDAD Y DOCUMENTACIÓN A LLENAR",
        font=("Helvetica", 16, "bold"),
        text_color="#1F4E79",
    ).pack(anchor="w", padx=15, pady=12)

    col_lista = df_a.columns[1] if len(df_a.columns) > 1 else col_mat_a
    for _, row in df_a.iterrows():
      item_txt = limpiar_texto(row[col_lista])
      if item_txt:
        card_item = ctk.CTkFrame(
            frame_a,
            fg_color="#FFFFFF",
            border_width=1,
            border_color="#E0E0E0",
            corner_radius=6,
        )
        card_item.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(
            card_item,
            text=f"• {item_txt}",
            font=("Helvetica", 15, "bold"),
            text_color="#111111",
            anchor="w",
        ).pack(padx=14, pady=12, fill="x")

  # ==============================================================================
  # PESTAÑA 5: PASOS OPERATIVOS
  # ==============================================================================
  def _setup_tab_pasos(self):
    self.frame_paso_container = ctk.CTkFrame(self.tab_pasos, corner_radius=10)
    self.frame_paso_container.pack(fill="both", expand=True, padx=5, pady=5)

    self.lbl_paso_num = ctk.CTkLabel(
        self.frame_paso_container,
        text="",
        font=("Helvetica", 18, "bold"),
        text_color="#1F4E79",
    )
    self.lbl_paso_num.pack(pady=(10, 5))

    self.frame_contenido_paso = ctk.CTkFrame(
        self.frame_paso_container, fg_color="transparent"
    )
    self.frame_contenido_paso.pack(fill="both", expand=True, padx=15, pady=5)

    self.frame_col_izq = ctk.CTkFrame(
        self.frame_contenido_paso, width=450, fg_color="transparent"
    )
    self.frame_col_izq.pack(side="left", fill="both", padx=(0, 10))

    self.box_paso_desc = ctk.CTkTextbox(
        self.frame_col_izq, font=("Helvetica", 15), corner_radius=6
    )
    self.box_paso_desc.pack(fill="both", expand=True, pady=(0, 10))

    self.lbl_paso_nota = ctk.CTkLabel(
        self.frame_col_izq,
        text="",
        font=("Helvetica", 14, "bold"),
        text_color="#B71C1C",
        anchor="w",
        justify="left",
        wraplength=420,
    )
    self.lbl_paso_nota.pack(fill="x", pady=2)

    self.frame_col_der = ctk.CTkFrame(
        self.frame_contenido_paso,
        fg_color="#FFFFFF",
        corner_radius=8,
        border_width=1,
        border_color="#CCCCCC",
    )
    self.frame_col_der.pack(side="right", fill="both", expand=True)

    self.lbl_paso_imagen = ctk.CTkLabel(
        self.frame_col_der, text="Sin imagen de paso"
    )
    self.lbl_paso_imagen.pack(expand=True, fill="both", padx=10, pady=10)

    self.frame_nav_botones = ctk.CTkFrame(
        self.frame_paso_container, fg_color="transparent"
    )
    self.frame_nav_botones.pack(pady=8)

    self.btn_anterior = ctk.CTkButton(
        self.frame_nav_botones,
        text="◄ Paso Anterior",
        font=("Helvetica", 13, "bold"),
        command=self.paso_anterior,
        width=150,
    )
    self.btn_anterior.pack(side="left", padx=10)

    self.btn_siguiente = ctk.CTkButton(
        self.frame_nav_botones,
        text="Paso Siguiente ►",
        font=("Helvetica", 13, "bold"),
        command=self.paso_siguiente,
        width=150,
    )
    self.btn_siguiente.pack(side="left", padx=10)

    # Leyenda de Puntos de Seguridad
    self.frame_puntos_seguridad = ctk.CTkFrame(
        self.tab_pasos,
        height=120,
        corner_radius=8,
        fg_color="#F5F5F5",
        border_width=1,
        border_color="#CCCCCC",
    )
    self.frame_puntos_seguridad.pack(fill="x", padx=5, pady=(5, 0))

    lbl_tit_seg = ctk.CTkLabel(
        self.frame_puntos_seguridad,
        text="Puntos de Seguridad",
        font=("Helvetica", 13, "bold"),
        fg_color="#FFC107",
        text_color="#000000",
    )
    lbl_tit_seg.pack(fill="x")

    self.grid_leyenda = ctk.CTkFrame(
        self.frame_puntos_seguridad, fg_color="transparent"
    )
    self.grid_leyenda.pack(fill="both", expand=True, padx=4, pady=4)

    for c in range(4):
      self.grid_leyenda.grid_columnconfigure(c, weight=1)

  def renderizar_tab_pasos(self, material):
    col_mat_p = self.data_pasos.columns[0]
    self.pasos_filtrados = (
        self.data_pasos[
            self.data_pasos[col_mat_p].apply(limpiar_texto) == material
        ].to_dict("records")
        if not self.data_pasos.empty
        else []
    )

    self.paso_actual_index = 0
    self.actualizar_vista_paso()
    self.renderizar_leyenda_seguridad()

  def actualizar_vista_paso(self):
    total = len(self.pasos_filtrados)
    if total == 0:
      self.lbl_paso_num.configure(text="SOP SIN PASOS REGISTRADOS")
      self.box_paso_desc.delete("1.0", "end")
      self.lbl_paso_nota.configure(text="")
      self.lbl_paso_imagen.configure(image=None, text="Sin imagen de paso")
      self.btn_anterior.configure(state="disabled")
      self.btn_siguiente.configure(state="disabled")
      return

    paso_data = self.pasos_filtrados[self.paso_actual_index]
    nombre_paso = limpiar_texto(paso_data.get(self.data_pasos.columns[1]))
    desc_paso = limpiar_texto(paso_data.get(self.data_pasos.columns[2]))
    nota_paso = limpiar_texto(paso_data.get(self.data_pasos.columns[3]))
    simb_paso = limpiar_texto(paso_data.get(self.data_pasos.columns[4]))

    self.lbl_paso_num.configure(text=nombre_paso.upper())

    self.box_paso_desc.delete("1.0", "end")
    self.box_paso_desc.insert("1.0", desc_paso)

    if nota_paso:
      txt_nota = f"[{simb_paso}] {nota_paso}" if simb_paso else nota_paso
      self.lbl_paso_nota.configure(text=txt_nota)
    else:
      self.lbl_paso_nota.configure(text="")

    imgs_p = self.imgs_pasos.get(self.material_actual, [])
    if self.paso_actual_index < len(imgs_p):
      pil_img = imgs_p[self.paso_actual_index]
      ctk_img = ctk.CTkImage(
          light_image=pil_img, dark_image=pil_img, size=(480, 340)
      )
      self.lbl_paso_imagen.configure(image=ctk_img, text="")
    else:
      self.lbl_paso_imagen.configure(image=None, text="Sin imagen disponible")

    self.btn_anterior.configure(
        state="normal" if self.paso_actual_index > 0 else "disabled"
    )
    self.btn_siguiente.configure(
        state="normal" if self.paso_actual_index < total - 1 else "disabled"
    )

  def renderizar_leyenda_seguridad(self):
    for child in self.grid_leyenda.winfo_children():
      child.destroy()

    if self.data_puntos_seg.empty:
      return

    filas_seg = self.data_puntos_seg.to_dict("records")
    colors_hdr = ["#FFF176", "#81C784", "#FF8A65", "#E0E0E0"]

    for c in range(min(4, len(filas_seg))):
      r_data = filas_seg[c]
      tit = limpiar_texto(r_data.get(self.data_puntos_seg.columns[0]))
      desc = limpiar_texto(r_data.get(self.data_puntos_seg.columns[1]))
      simb = limpiar_texto(r_data.get(self.data_puntos_seg.columns[2]))

      cell = ctk.CTkFrame(
          self.grid_leyenda,
          border_width=1,
          border_color="#A0A0A0",
          corner_radius=2,
          fg_color="#FFFFFF",
      )
      cell.grid(row=0, column=c, sticky="nsew", padx=2, pady=1)

      hdr = ctk.CTkFrame(
          cell, fg_color=colors_hdr[c % 4], corner_radius=0, height=22
      )
      hdr.pack(fill="x")

      ctk.CTkLabel(
          hdr, text=tit, font=("Helvetica", 10, "bold"), text_color="#000000"
      ).pack(pady=2)

      body = ctk.CTkFrame(cell, fg_color="transparent")
      body.pack(fill="both", expand=True, padx=4, pady=2)

      ctk.CTkLabel(
          body,
          text=desc,
          font=("Helvetica", 9),
          text_color="#000000",
          justify="left",
          wraplength=210,
          anchor="w",
      ).pack(side="left", fill="both", expand=True)

      if simb:
        lbl_s = ctk.CTkLabel(
            body,
            text=f"[{simb}]",
            font=("Helvetica", 12, "bold"),
            text_color="#1F4E79",
            width=35,
        )
        lbl_s.pack(side="right", padx=2)

  def paso_anterior(self):
    if self.paso_actual_index > 0:
      self.paso_actual_index -= 1
      self.actualizar_vista_paso()

  def paso_siguiente(self):
    if self.paso_actual_index < len(self.pasos_filtrados) - 1:
      self.paso_actual_index += 1
      self.actualizar_vista_paso()

  # ==============================================================================
  # PESTAÑA 6: HISTORIAL
  # ==============================================================================
  def renderizar_tab_historial(self, material):
    for child in self.tab_historial.winfo_children():
      child.destroy()

    scroll = ctk.CTkScrollableFrame(self.tab_historial)
    scroll.pack(fill="both", expand=True)

    col_mat_h = self.data_historial.columns[0]
    df_h = self.data_historial[
        self.data_historial[col_mat_h].apply(limpiar_texto) == material
    ]

    if df_h.empty:
      ctk.CTkLabel(
          scroll,
          text="No hay historial de cambios registrado.",
          font=("Helvetica", 15, "bold"),
      ).pack(pady=40)
      return

    frame_h = ctk.CTkFrame(scroll, corner_radius=8)
    frame_h.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        frame_h,
        text="📜 HISTORIAL DE CAMBIOS Y REVISIONES",
        font=("Helvetica", 16, "bold"),
        text_color="#1F4E79",
    ).pack(anchor="w", padx=15, pady=12)

    for _, row in df_h.iterrows():
      rev = limpiar_texto(row.iloc[1]) if len(row) > 1 else ""
      fecha = limpiar_texto(row.iloc[2]) if len(row) > 2 else ""
      depto = limpiar_texto(row.iloc[3]) if len(row) > 3 else ""
      desc = limpiar_texto(row.iloc[4]) if len(row) > 4 else ""

      card = ctk.CTkFrame(
          frame_h,
          fg_color="#FFFFFF",
          border_width=1,
          border_color="#E0E0E0",
          corner_radius=6,
      )
      card.pack(fill="x", padx=15, pady=4)

      lbl_t = f"Rev: {rev}  │  Fecha: {fecha}  │  Depto: {depto}"
      ctk.CTkLabel(
          card,
          text=lbl_t,
          font=("Helvetica", 14, "bold"),
          text_color="#1F4E79",
          anchor="w",
      ).pack(padx=14, pady=(10, 2), fill="x")
      ctk.CTkLabel(
          card,
          text=desc,
          font=("Helvetica", 14),
          text_color="#333333",
          anchor="w",
          justify="left",
      ).pack(padx=14, pady=(0, 10), fill="x")


if __name__ == "__main__":
  app = SOPApp()
  app.mainloop()
